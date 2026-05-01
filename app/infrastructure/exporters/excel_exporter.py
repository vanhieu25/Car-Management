"""ExcelExporter - Excel export utility using openpyxl.

Provides generic Excel export with formatting for financial reports.
Features:
- Header bold (first row)
- Freeze pane (first row + first column header)
- Auto-fit column width
- Support multiple sheets via sheet_config list
- Handle money formatting (VND with thousand separator)
- Handle date formatting
"""

import os
from dataclasses import dataclass
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


class ExcelExporterError(Exception):
    """Raised when Excel export fails."""
    pass


class ValidationError(ExcelExporterError):
    """Raised when validation fails."""
    pass


# Column format types
FORMAT_MONEY = "money"
FORMAT_DATE = "date"
FORMAT_DATETIME = "datetime"
FORMAT_PERCENT = "percent"
FORMAT_NUMBER = "number"
FORMAT_TEXT = "text"


@dataclass
class ColumnConfig:
    """Configuration for a single column."""
    header: str  # Column header text
    key: str  # Key in the data dict
    width: int = 15  # Column width (auto-calculated if 0)
    format: str = FORMAT_TEXT  # Format type
    bold: bool = False  # Override header bold


@dataclass
class SheetConfig:
    """Configuration for a single sheet."""
    name: str  # Sheet tab name
    columns: List[ColumnConfig]  # Column definitions
    title: Optional[str] = None  # Optional sheet title


class ExcelExporter:
    """Utility class for exporting data to Excel with formatting."""

    VND_FORMAT = '#,##0'

    def __init__(self):
        """Initialize ExcelExporter."""
        self._wb = None

    def export_report(
        self,
        report_data: List[Dict[str, Any]],
        sheet_config: Dict[str, Any],
        output_path: str,
    ) -> str:
        """Export report data to Excel file.

        Args:
            report_data: List of dicts (rows) to export.
            sheet_config: Dict with:
                - name: sheet tab name
                - columns: list of ColumnConfig or dict with header/key/width/format
                - title: optional sheet title
            output_path: Full path to save .xlsx file.

        Returns:
            Path to the saved Excel file.

        Raises:
            ValidationError: If config is invalid.
            ExcelExporterError: If write fails.
        """
        try:
            self._wb = Workbook()
            ws = self._wb.active
            ws.title = sheet_config.get("name", "Sheet1")

            # Handle columns config
            columns = sheet_config.get("columns", [])
            if not columns:
                raise ValidationError("columns config is required")

            # Parse columns to ColumnConfig objects
            col_configs = []
            for col in columns:
                if isinstance(col, dict):
                    col_configs.append(ColumnConfig(
                        header=col.get("header", ""),
                        key=col.get("key", ""),
                        width=col.get("width", 15),
                        format=col.get("format", FORMAT_TEXT),
                        bold=col.get("bold", False),
                    ))
                elif isinstance(col, ColumnConfig):
                    col_configs.append(col)
                else:
                    raise ValidationError(f"Invalid column config: {col}")

            # Add title row if specified
            title_row = 1
            if sheet_config.get("title"):
                ws.merge_cells(
                    start_row=1, start_column=1,
                    end_row=1, end_column=len(col_configs)
                )
                cell = ws.cell(row=1, column=1, value=sheet_config["title"])
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal="center")
                title_row = 2

            # Write header row
            header_row = title_row
            for col_idx, col_config in enumerate(col_configs, start=1):
                cell = ws.cell(row=header_row, column=col_idx, value=col_config.header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Light gray background for header
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="D9D9D9"
                )
                thin_border = Border(
                    bottom=Side(style='thin')
                )
                cell.border = thin_border

            # Write data rows
            for row_idx, row_data in enumerate(report_data, start=header_row + 1):
                for col_idx, col_config in enumerate(col_configs, start=1):
                    value = row_data.get(col_config.key, "")
                    cell = ws.cell(row=row_idx, column=col_idx)

                    # Format value based on column format
                    if col_config.format == FORMAT_MONEY and value:
                        cell.value = value
                        cell.number_format = self.VND_FORMAT
                    elif col_config.format == FORMAT_DATE and value:
                        cell.value = str(value)[:10] if value else ""
                    elif col_config.format == FORMAT_DATETIME and value:
                        cell.value = str(value) if value else ""
                    elif col_config.format == FORMAT_PERCENT and value:
                        cell.value = value / 100 if isinstance(value, (int, float)) else value
                        cell.number_format = "0.00%"
                    elif col_config.format == FORMAT_NUMBER and value:
                        cell.value = value
                        cell.number_format = "#,##0"
                    else:
                        cell.value = value if value is not None else ""

                    cell.alignment = Alignment(horizontal="left")

            # Auto-fit column widths
            for col_idx, col_config in enumerate(col_configs, start=1):
                max_length = len(col_config.header)
                for row_idx in range(header_row + 1, header_row + len(report_data) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                # Set width with padding
                adjusted_width = min(max(max_length + 2, col_config.width), 40)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            # Freeze pane (below header, after title)
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

            # Ensure directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # Save
            self._wb.save(output_path)
            self._wb = None

            return output_path

        except ValidationError:
            raise
        except Exception as e:
            self._wb = None
            raise ExcelExporterError(f"Failed to export Excel: {str(e)}")

    def export_multi_sheet(
        self,
        sheets: List[Dict[str, Any]],
        output_path: str,
    ) -> str:
        """Export multiple sheets to a single Excel file.

        Args:
            sheets: List of sheet configs (each is a dict with name/columns/title/data).
            output_path: Full path to save .xlsx file.

        Returns:
            Path to the saved Excel file.
        """
        try:
            self._wb = Workbook()
            # Remove default sheet
            if self._wb.worksheets:
                self._wb.remove(self._wb.worksheets[0])

            for sheet_data in sheets:
                ws = self._wb.create_sheet(title=sheet_data.get("name", "Sheet"))

                columns = sheet_data.get("columns", [])
                if not columns:
                    continue

                # Parse columns
                col_configs = []
                for col in columns:
                    if isinstance(col, dict):
                        col_configs.append(ColumnConfig(
                            header=col.get("header", ""),
                            key=col.get("key", ""),
                            width=col.get("width", 15),
                            format=col.get("format", FORMAT_TEXT),
                            bold=col.get("bold", False),
                        ))
                    else:
                        col_configs.append(col)

                # Add title row if specified
                title_row = 1
                if sheet_data.get("title"):
                    ws.merge_cells(
                        start_row=1, start_column=1,
                        end_row=1, end_column=len(col_configs)
                    )
                    cell = ws.cell(row=1, column=1, value=sheet_data["title"])
                    cell.font = Font(bold=True, size=14)
                    cell.alignment = Alignment(horizontal="center")
                    title_row = 2

                # Write header row
                header_row = title_row
                for col_idx, col_config in enumerate(col_configs, start=1):
                    cell = ws.cell(row=header_row, column=col_idx, value=col_config.header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor="D9D9D9"
                    )
                    thin_border = Border(bottom=Side(style='thin'))
                    cell.border = thin_border

                # Write data rows
                report_data = sheet_data.get("data", [])
                for row_idx, row_data in enumerate(report_data, start=header_row + 1):
                    for col_idx, col_config in enumerate(col_configs, start=1):
                        value = row_data.get(col_config.key, "")
                        cell = ws.cell(row=row_idx, column=col_idx)

                        if col_config.format == FORMAT_MONEY and value:
                            cell.value = value
                            cell.number_format = self.VND_FORMAT
                        elif col_config.format == FORMAT_DATE and value:
                            cell.value = str(value)[:10] if value else ""
                        elif col_config.format == FORMAT_DATETIME and value:
                            cell.value = str(value) if value else ""
                        elif col_config.format == FORMAT_PERCENT and value:
                            cell.value = value / 100 if isinstance(value, (int, float)) else value
                            cell.number_format = "0.00%"
                        elif col_config.format == FORMAT_NUMBER and value:
                            cell.value = value
                            cell.number_format = "#,##0"
                        else:
                            cell.value = value if value is not None else ""

                        cell.alignment = Alignment(horizontal="left")

                # Auto-fit column widths
                for col_idx, col_config in enumerate(col_configs, start=1):
                    max_length = len(col_config.header)
                    for row_idx in range(header_row + 1, header_row + len(report_data) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = min(max(max_length + 2, col_config.width), 40)
                    ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

                # Freeze pane
                ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

            # Ensure directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            self._wb.save(output_path)
            self._wb = None

            return output_path

        except Exception as e:
            self._wb = None
            raise ExcelExporterError(f"Failed to export multi-sheet Excel: {str(e)}")