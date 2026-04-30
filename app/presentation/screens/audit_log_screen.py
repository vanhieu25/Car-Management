"""Audit log screen - S-SYS-01 - Audit log viewer.

Features:
- Table with columns: thời gian, user, hành động, bảng ảnh hưởng, ban_ghi_id
- Filters: ngày từ-đến, user, action, table
- Pagination: 50 rows per page
- Export to Excel button
- Row click → detail dialog (S-SYS-02)

References:
- UC-SEC-03: Audit log viewing
- BR-SEC-09: Audit trail requirement
"""

from datetime import datetime
from typing import Optional

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget,
    QTableWidgetItem, QPushButton, QLineEdit, QComboBox,
    QDateEdit, QDialog, QTextEdit, QHeaderView, QAbstractItemView,
    QApplication, QMessageBox
)
from PyQt6.QtCore import Qt, QDate, QTimer, pyqtSignal
from PyQt6.QtGui import QFont, QColor

from app.application.services.audit_log_service import AuditLogService, AuditLogEntry
from app.application.services.session import CurrentSession


# Pagination constants
PAGE_SIZE = 50


class AuditLogScreen(QWidget):
    """Audit log viewer screen - S-SYS-01."""
    
    def __init__(self, db_conn, session: CurrentSession, parent=None):
        """Initialize audit log screen.
        
        Args:
            db_conn: sqlite3 database connection.
            session: Current user session.
            parent: Parent widget.
        """
        super().__init__(parent)
        self._db_conn = db_conn
        self._session = session
        self._audit_service = AuditLogService(db_conn)
        
        self._current_page = 0
        self._total_count = 0
        self._filters = {
            "from_date": None,
            "to_date": None,
            "nhan_vien_id": None,
            "action": None,
            "table": None,
        }
        
        self._setup_ui()
        self._load_data()
    
    def _setup_ui(self):
        """Set up UI components."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)
        
        # Header
        header_layout = QHBoxLayout()
        title = QLabel("Nhật ký hoạt động")
        title.setStyleSheet("font-size: 24px; font-weight: 600; color: #1d1d1f;")
        header_layout.addWidget(title)
        header_layout.addStretch()
        
        # Export button
        self._export_btn = QPushButton("📥 Xuất Excel")
        self._export_btn.setStyleSheet("""
            QPushButton {
                background-color: #0066cc;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #0055aa;
            }
        """)
        self._export_btn.clicked.connect(self._on_export)
        header_layout.addWidget(self._export_btn)
        
        layout.addLayout(header_layout)
        
        # Filter section
        filter_group = QWidget()
        filter_group.setStyleSheet("""
            QWidget {
                background-color: #f5f5f7;
                border-radius: 8px;
                padding: 16px;
            }
        """)
        filter_layout = QHBoxLayout(filter_group)
        filter_layout.setContentsMargins(16, 16, 16, 16)
        filter_layout.setSpacing(12)
        
        # Date range filters
        filter_layout.addWidget(QLabel("Từ ngày:"))
        self._from_date_edit = QDateEdit()
        self._from_date_edit.setCalendarPopup(True)
        self._from_date_edit.setDate(QDate.currentDate().addDays(-7))
        self._from_date_edit.setDisplayFormat("yyyy-MM-dd")
        filter_layout.addWidget(self._from_date_edit)
        
        filter_layout.addWidget(QLabel("Đến ngày:"))
        self._to_date_edit = QDateEdit()
        self._to_date_edit.setCalendarPopup(True)
        self._to_date_edit.setDate(QDate.currentDate())
        self._to_date_edit.setDisplayFormat("yyyy-MM-dd")
        filter_layout.addWidget(self._to_date_edit)
        
        # Action filter
        filter_layout.addWidget(QLabel("Hành động:"))
        self._action_combo = QComboBox()
        self._action_combo.addItems(["Tất cả", "LOGIN", "LOGOUT", "CREATE_", "UPDATE_", "DELETE_", "CHANGE_PASSWORD", "CANCEL_HD"])
        filter_layout.addWidget(self._action_combo)
        
        # Table filter
        filter_layout.addWidget(QLabel("Bảng:"))
        self._table_combo = QComboBox()
        self._table_combo.addItems(["Tất cả", "xe", "khach_hang", "hop_dong", "nhan_vien", "phu_kien", "khuyen_mai", "bao_hanh"])
        filter_layout.addWidget(self._table_combo)
        
        # Search button
        self._search_btn = QPushButton("🔍 Tìm kiếm")
        self._search_btn.setStyleSheet("""
            QPushButton {
                background-color: #34c759;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #2db14e;
            }
        """)
        self._search_btn.clicked.connect(self._on_search)
        filter_layout.addWidget(self._search_btn)
        
        # Clear button
        self._clear_btn = QPushButton("🗑️ Xóa filter")
        self._clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff3b30;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #e0342c;
            }
        """)
        self._clear_btn.clicked.connect(self._on_clear_filters)
        filter_layout.addWidget(self._clear_btn)
        
        layout.addWidget(filter_group)
        
        # Table
        self._table = QTableWidget()
        self._table.setColumnCount(6)
        self._table.setHorizontalHeaderLabels([
            "ID", "Thời gian", "Người dùng", "Hành động", "Bảng", "Mã bản ghi"
        ])
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #d2d2d7;
                border-radius: 8px;
                gridline-color: #e5e5ea;
                background-color: white;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QHeaderView::section {
                background-color: #f5f5f7;
                padding: 8px;
                border: none;
                font-weight: 600;
            }
            QTableWidget::item:selected {
                background-color: #0066cc;
                color: white;
            }
        """)
        self._table.cellClicked.connect(self._on_row_clicked)
        layout.addWidget(self._table)
        
        # Pagination
        pagination_layout = QHBoxLayout()
        pagination_layout.addStretch()
        
        self._prev_btn = QPushButton("◀ Trước")
        self._prev_btn.clicked.connect(self._on_prev_page)
        pagination_layout.addWidget(self._prev_btn)
        
        self._page_label = QLabel("Trang 1 / 1")
        self._page_label.setStyleSheet("font-size: 14px; color: #86868b;")
        pagination_layout.addWidget(self._page_label)
        
        self._next_btn = QPushButton("Sau ▶")
        self._next_btn.clicked.connect(self._on_next_page)
        pagination_layout.addWidget(self._next_btn)
        
        self._total_label = QLabel("Tổng: 0 bản ghi")
        self._total_label.setStyleSheet("font-size: 14px; color: #86868b; margin-left: 16px;")
        pagination_layout.addWidget(self._total_label)
        
        layout.addLayout(pagination_layout)
    
    def _load_data(self):
        """Load audit log data based on filters and pagination."""
        # Build filter values
        from_date = self._from_date_edit.date().toPython().isoformat() if self._from_date_edit.date().isValid() else None
        to_date = self._to_date_edit.date().toPython().isoformat() if self._to_date_edit.date().isValid() else None
        action_filter = self._action_combo.currentText() if self._action_combo.currentText() != "Tất cả" else None
        table_filter = self._table_combo.currentText() if self._table_combo.currentText() != "Tất cả" else None
        
        # Count total
        self._total_count = self._audit_service.count(
            nhan_vien_id=None,
            action=action_filter,
            table=table_filter,
            from_time=datetime.fromisoformat(from_date) if from_date else None,
            to_time=datetime.fromisoformat(to_date) if to_date else None,
        )
        
        # Calculate total pages
        total_pages = max(1, (self._total_count + PAGE_SIZE - 1) // PAGE_SIZE)
        if self._current_page >= total_pages:
            self._current_page = total_pages - 1
        
        # Load data
        offset = self._current_page * PAGE_SIZE
        entries = self._audit_service.query(
            nhan_vien_id=None,
            action=action_filter,
            table=table_filter,
            from_time=datetime.fromisoformat(from_date) if from_date else None,
            to_time=datetime.fromisoformat(to_date) if to_date else None,
            limit=PAGE_SIZE,
            offset=offset,
        )
        
        # Populate table
        self._table.setRowCount(len(entries))
        for row, entry in enumerate(entries):
            self._table.setItem(row, 0, QTableWidgetItem(str(entry["id"])))
            self._table.setItem(row, 1, QTableWidgetItem(entry["thoi_gian"][:19].replace("T", " ")))
            self._table.setItem(row, 2, QTableWidgetItem(entry.get("ho_ten") or entry.get("username", "N/A")))
            self._table.setItem(row, 3, QTableWidgetItem(entry["hanh_dong"]))
            self._table.setItem(row, 4, QTableWidgetItem(entry["bang_anh_huong"] or "-"))
            self._table.setItem(row, 5, QTableWidgetItem(str(entry["ban_ghi_id"]) if entry["ban_ghi_id"] else "-"))
            
            # Store full entry data for detail dialog
            self._table.item(row, 0).setData(Qt.ItemDataRole.UserRole, entry)
        
        # Update pagination UI
        self._page_label.setText(f"Trang {self._current_page + 1} / {total_pages}")
        self._total_label.setText(f"Tổng: {self._total_count} bản ghi")
        
        # Enable/disable nav buttons
        self._prev_btn.setEnabled(self._current_page > 0)
        self._next_btn.setEnabled((self._current_page + 1) * PAGE_SIZE < self._total_count)
    
    def _on_search(self):
        """Handle search button click."""
        self._current_page = 0
        self._load_data()
    
    def _on_clear_filters(self):
        """Clear all filters and reload."""
        self._from_date_edit.setDate(QDate.currentDate().addDays(-7))
        self._to_date_edit.setDate(QDate.currentDate())
        self._action_combo.setCurrentIndex(0)
        self._table_combo.setCurrentIndex(0)
        self._current_page = 0
        self._load_data()
    
    def _on_prev_page(self):
        """Go to previous page."""
        if self._current_page > 0:
            self._current_page -= 1
            self._load_data()
    
    def _on_next_page(self):
        """Go to next page."""
        total_pages = max(1, (self._total_count + PAGE_SIZE - 1) // PAGE_SIZE)
        if self._current_page < total_pages - 1:
            self._current_page += 1
            self._load_data()
    
    def _on_row_clicked(self, row, column):
        """Handle row click - show detail dialog."""
        item = self._table.item(row, 0)
        if item:
            entry = item.data(Qt.ItemDataRole.UserRole)
            if entry:
                dialog = AuditLogDetailDialog(entry, self)
                dialog.exec()
    
    def _on_export(self):
        """Export filtered audit log to Excel."""
        from PyQt6.QtWidgets import QFileDialog
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Xuất nhật ký hoạt động",
            f"audit_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Audit Log"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ["ID", "Thời gian", "Người dùng", "Hành động", "Bảng", "Mã bản ghi", "Nội dung JSON"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Get all filtered data (not just current page)
            from_date = self._from_date_edit.date().toPython().isoformat() if self._from_date_edit.date().isValid() else None
            to_date = self._to_date_edit.date().toPython().isoformat() if self._to_date_edit.date().isValid() else None
            action_filter = self._action_combo.currentText() if self._action_combo.currentText() != "Tất cả" else None
            table_filter = self._table_combo.currentText() if self._table_combo.currentText() != "Tất cả" else None
            
            entries = self._audit_service.query(
                nhan_vien_id=None,
                action=action_filter,
                table=table_filter,
                from_time=datetime.fromisoformat(from_date) if from_date else None,
                to_time=datetime.fromisoformat(to_date) if to_date else None,
                limit=10000,  # Get all
                offset=0,
            )
            
            # Data rows
            for row_idx, entry in enumerate(entries, 2):
                row_data = [
                    entry["id"],
                    entry["thoi_gian"][:19].replace("T", " "),
                    entry.get("ho_ten") or entry.get("username", "N/A"),
                    entry["hanh_dong"],
                    entry["bang_anh_huong"] or "-",
                    str(entry["ban_ghi_id"]) if entry["ban_ghi_id"] else "-",
                    str(entry.get("noi_dung", {}))
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
            
            # Auto column width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "Thành công",
                f"Đã xuất {len(entries)} bản ghi ra file Excel:\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Lỗi",
                f"Không thể xuất Excel: {str(e)}"
            )


class AuditLogDetailDialog(QDialog):
    """Dialog to display detailed audit log entry with JSON diff.
    
    Shows before/after JSON in formatted, syntax-highlighted style.
    """
    
    def __init__(self, entry: dict, parent=None):
        """Initialize detail dialog.
        
        Args:
            entry: Audit log entry dict.
            parent: Parent widget.
        """
        super().__init__(parent)
        self._entry = entry
        self._setup_ui()
    
    def _setup_ui(self):
        """Set up dialog UI."""
        self.setWindowTitle(f"Nhật ký #{self._entry['id']}")
        self.setMinimumSize(700, 500)
        self.setStyleSheet("""
            QDialog {
                background-color: #ffffff;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)
        
        # Header info
        header_widget = QWidget()
        header_widget.setStyleSheet("""
            QWidget {
                background-color: #f5f5f7;
                border-radius: 8px;
                padding: 16px;
            }
        """)
        header_layout = QVBoxLayout(header_widget)
        
        info_text = f"""
<b>Hành động:</b> {self._entry['hanh_dong']}
<b>Bảng ảnh hưởng:</b> {self._entry.get('bang_anh_huong', 'N/A')}
<b>Mã bản ghi:</b> {self._entry.get('ban_ghi_id', 'N/A')}
<b>Người thực hiện:</b> {self._entry.get('ho_ten', 'N/A')} ({self._entry.get('username', 'N/A')})
<b>Thời gian:</b> {self._entry.get('thoi_gian', 'N/A')}
        """
        header_label = QLabel(info_text.strip())
        header_label.setStyleSheet("font-size: 14px; color: #1d1d1f;")
        header_layout.addWidget(header_label)
        
        layout.addWidget(header_widget)
        
        # JSON diff section
        layout.addWidget(QLabel("Nội dung chi tiết (JSON):"))
        
        json_display = QTextEdit()
        json_display.setReadOnly(True)
        json_display.setStyleSheet("""
            QTextEdit {
                background-color: #1e1e1e;
                color: #d4d4d4;
                border: 1px solid #d2d2d7;
                border-radius: 8px;
                padding: 12px;
                font-family: 'SF Mono', 'Monaco', 'Consolas', monospace;
                font-size: 13px;
            }
        """)
        
        # Format JSON content
        noi_dung = self._entry.get("noi_dung", {})
        
        if isinstance(noi_dung, str):
            try:
                import json
                noi_dung = json.loads(noi_dung)
            except:
                pass
        
        # Build formatted display
        formatted_lines = []
        
        if "before" in noi_dung and noi_dung["before"]:
            formatted_lines.append("┌─ BEFORE ──────────────────────────────")
            before_str = self._format_dict(noi_dung["before"])
            formatted_lines.append(before_str)
            formatted_lines.append("└────────────────────────────────────────")
        
        if "after" in noi_dung and noi_dung["after"]:
            formatted_lines.append("┌─ AFTER ───────────────────────────────")
            after_str = self._format_dict(noi_dung["after"])
            formatted_lines.append(after_str)
            formatted_lines.append("└────────────────────────────────────────")
        
        if "changes" in noi_dung and noi_dung["changes"]:
            formatted_lines.append("┌─ CHANGES ──────────────────────────────")
            changes_str = self._format_dict(noi_dung["changes"])
            formatted_lines.append(changes_str)
            formatted_lines.append("└────────────────────────────────────────")
        
        if not formatted_lines:
            formatted_lines.append("(Không có chi tiết bổ sung)")
        
        json_display.setPlainText("\n".join(formatted_lines))
        layout.addWidget(json_display, stretch=1)
        
        # Close button
        close_btn = QPushButton("Đóng")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #86868b;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 24px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #636366;
            }
        """)
        close_btn.clicked.connect(self.accept)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)
    
    def _format_dict(self, d: dict, indent: int = 0) -> str:
        """Format a dict for display with syntax highlighting colors.
        
        Args:
            d: Dict to format.
            indent: Indentation level.
            
        Returns:
            Formatted string.
        """
        if not d:
            return "  (empty)"
        
        lines = []
        prefix = "  " * indent
        
        for key, value in d.items():
            if isinstance(value, dict):
                lines.append(f"{prefix}<span style='color:#9cdcfe;'>{key}</span>:")
                lines.append(self._format_dict(value, indent + 1))
            elif isinstance(value, list):
                lines.append(f"{prefix}<span style='color:#9cdcfe;'>{key}</span>: [{len(value)} items]")
            elif value is None:
                lines.append(f"{prefix}<span style='color:#9cdcfe;'>{key}</span>: <span style='color:#569cd6;'>null</span>")
            elif isinstance(value, bool):
                color = "#569cd6" if value else "#ce9178"
                lines.append(f"{prefix}<span style='color:#9cdcfe;'>{key}</span>: <span style='color:{color};'>{value}</span>")
            elif isinstance(value, (int, float)):
                lines.append(f"{prefix}<span style='color:#b5cea8;'>{key}</span>: <span style='color:#b5cea8;'>{value}</span>")
            else:
                lines.append(f"{prefix}<span style='color:#9cdcfe;'>{key}</span>: <span style='color:#ce9178;'>\"{value}\"</span>")
        
        return "\n".join(lines) if lines else f"{prefix}(empty)"
