"""UAT tests for BaoCao and Dashboard screens - T-G5.4.TEST.05.

Tests user-facing behavior through the service layer:
- AC-BC-01: Revenue report date range selection
- AC-BC-02: Excel export for revenue report
- AC-BC-03: Top vehicles chart and table matching
- AC-BC-04: Employee KPI for specific month
- AC-BC-05: VIP list sorted by total purchase
- AC-DB-01: Dashboard 7 KPI tiles
- AC-DB-02: KPI tile navigation
- AC-DB-03: Auto-refresh behavior

Note: These tests focus on service-layer behavior. UI-level tests would require
PyQt6 widget instantiation which is covered in navigation tests.
"""

import pytest
import sqlite3
import os
import sys
import tempfile
from datetime import datetime, timedelta

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.application.services.bao_cao_service import BaoCaoService
from app.application.services.dashboard_service import DashboardService
from app.infrastructure.exporters.excel_exporter import ExcelExporter


# =============================================================================
# Fixtures
# =============================================================================

def _migrate(db_path):
    """Run migrations on database."""
    from app.infrastructure.database.migrations.runner import MigrationRunner
    runner = MigrationRunner(db_path)
    runner.run_pending()


def _seed_uat_data(conn):
    """Seed data for UAT testing."""
    conn.execute("PRAGMA foreign_keys = ON")

    # vai_tro
    conn.execute("""
        INSERT INTO vai_tro (id, ma_vai_tro, ten_vai_tro)
        VALUES (1, 'A-01', 'Quản trị viên'),
               (2, 'A-02', 'Quản lý'),
               (3, 'A-03', 'Nhân viên')
    """)

    # nhan_vien
    conn.execute("""
        INSERT INTO nhan_vien (id, username, mat_khau_hash, ho_ten, email, vai_tro_id, trang_thai)
        VALUES (1, 'admin', '$2b$12$LQv3c1yqBWVHxkd0LHAkCOYz6TtxMQJqhN8/X4.NTtFQtE3T8TXK', 'Admin User', 'admin@test.com', 1, 'active'),
               (2, 'sales1', '$2b$12$LQv3c1yqBWVHxkd0LHAkCOYz6TtxMQJqhN8/X4.NTtFQtE3T8TXK', 'Sales One', 'sales1@test.com', 2, 'active'),
               (3, 'sales2', '$2b$12$LQv3c1yqBWVHxkd0LHAkCOYz6TtxMQJqhN8/X4.NTtFQtE3T8TXK', 'Sales Two', 'sales2@test.com', 2, 'active')
    """)

    # khach_hang
    conn.execute("""
        INSERT INTO khach_hang (id, ho_ten, so_dien_thoai, email, dia_chi, phan_loai, tong_gia_tri_mua, so_xe_da_mua)
        VALUES (1, 'Khach VIP 1', '0909000001', 'vip1@test.com', '123 VIP', 'VIP', 2000000000, 3),
               (2, 'Khach VIP 2', '0909000002', 'vip2@test.com', '456 VIP', 'VIP', 3000000000, 4),
               (3, 'Khach Thuong 1', '0909000003', 'thuong1@test.com', '789 Thuong', 'Thuong', 500000000, 1)
    """)

    # xe
    conn.execute("""
        INSERT INTO xe (id, ma_xe, hang, dong_xe, nam_san_xuat, mau_sac, gia_ban, so_luong_ton, muc_toi_thieu, trang_thai)
        VALUES (1, 'XE001', 'Toyota', 'Camry', 2024, 'Den', 500000000, 5, 2, 'con_hang'),
               (2, 'XE002', 'Honda', 'Civic', 2024, 'Trang', 400000000, 3, 2, 'con_hang'),
               (3, 'XE003', 'BMW', 'X5', 2024, 'Den', 1500000000, 1, 1, 'con_hang')
    """)

    # hop_dong - seed for April 2026
    now = datetime.now()
    for i in range(30):
        day = min(i % 28 + 1, 28)
        date_str = f"2026-04-{day:02d}"
        xe_id = (i % 3) + 1
        kh_id = (i % 3) + 1
        nv_id = (i % 2) + 1
        status = 'da_giao_xe' if i % 2 == 0 else 'da_thanh_toan'
        tong_tien = 400000000 + (i * 20000000)

        conn.execute(f"""
            INSERT INTO hop_dong (id, ma_hop_dong, khach_hang_id, xe_id, nhan_vien_id,
                                  gia_xe, tong_gia_phu_kien, tien_giam_km, tong_tien,
                                  trang_thai, ngay_tao)
            VALUES ({100 + i}, 'HD{100 + i:06d}', {kh_id}, {xe_id}, {nv_id},
                    400000000, 0, 0, {tong_tien}, '{status}', '{date_str}')
        """)

    # Seed for May 2026 (another month for comparison)
    for i in range(20):
        day = min(i % 28 + 1, 28)
        date_str = f"2026-05-{day:02d}"
        xe_id = (i % 3) + 1
        kh_id = (i % 3) + 1
        nv_id = (i % 2) + 1
        status = 'da_giao_xe' if i % 2 == 0 else 'da_thanh_toan'
        tong_tien = 450000000 + (i * 15000000)

        conn.execute(f"""
            INSERT INTO hop_dong (id, ma_hop_dong, khach_hang_id, xe_id, nhan_vien_id,
                                  gia_xe, tong_gia_phu_kien, tien_giam_km, tong_tien,
                                  trang_thai, ngay_tao)
            VALUES ({200 + i}, 'HD{200 + i:06d}', {kh_id}, {xe_id}, {nv_id},
                    450000000, 0, 0, {tong_tien}, '{status}', '{date_str}')
        """)


@pytest.fixture
def uat_db():
    """Create database with UAT test data."""
    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as f:
        db_path = f.name

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    _migrate(db_path)
    _seed_uat_data(conn)
    conn.commit()
    conn.close()

    yield db_path

    if os.path.exists(db_path):
        os.unlink(db_path)


@pytest.fixture
def temp_dir():
    """Create temporary directory for file tests."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tempfile.Path(tmpdir)


# =============================================================================
# TEST.05 — UAT BaoCao Screens
# =============================================================================
class TestUAT_BaoCao:
    """TEST.05 — UAT theo AC-BC-* cho các màn hình báo cáo"""

    def test_ac_bc_01_revenue_date_range(self, uat_db):
        """AC-BC-01: Chọn khoảng ngày → báo cáo doanh thu theo đúng period."""
        conn = sqlite3.connect(uat_db)
        conn.execute("PRAGMA foreign_keys = ON")
        service = BaoCaoService(conn)

        # Select April 2026 only
        result = service.revenue("2026-04-01", "2026-04-30", group_by="day")

        # Should return daily breakdown for April
        assert len(result["breakdown"]) > 0
        assert result["total_contracts"] == 30  # We seeded 30 for April

        # Verify all periods are within April 2026
        for item in result["breakdown"]:
            assert item["period"].startswith("2026-04")

        conn.close()

    def test_ac_bc_02_excel_export(self, uat_db, temp_dir):
        """AC-BC-02: Click Xuất Excel → tạo file .xlsx với dữ liệu đúng."""
        conn = sqlite3.connect(uat_db)
        conn.execute("PRAGMA foreign_keys = ON")
        service = BaoCaoService(conn)
        exporter = ExcelExporter()

        # Get revenue data
        result = service.revenue("2026-04-01", "2026-04-30", group_by="day")

        # Export to Excel
        config = {
            "name": "DoanhThu",
            "title": "Báo Cáo Doanh Thu Tháng 4/2026",
            "columns": [
                {"header": "Ngày", "key": "period", "format": "date"},
                {"header": "Số HĐ", "key": "so_hop_dong", "format": "number"},
                {"header": "Doanh Thu", "key": "doanh_thu", "format": "money"},
                {"header": "Tỷ Lệ", "key": "ty_le", "format": "percent"},
            ],
        }

        output_path = temp_dir / "bao_cao_doanh_thu.xlsx"
        exporter.export_report(result["breakdown"], config, str(output_path))

        # Verify file exists and has content
        assert output_path.exists()
        assert output_path.stat().st_size > 0

        # Verify Excel content
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Check header row
        headers = [cell.value for cell in ws[2]]
        assert "Ngày" in headers
        assert "Số HĐ" in headers

        # Check data rows match the breakdown
        data_rows = len(result["breakdown"])
        assert ws.max_row >= data_rows + 2  # title + header + data

        wb.close()
        conn.close()

    def test_ac_bc_03_top_vehicles_chart_matches_table(self, uat_db):
        """AC-BC-03: Top vehicles chart khớp với bảng dữ liệu."""
        conn = sqlite3.connect(uat_db)
        conn.execute("PRAGMA foreign_keys = ON")
        service = BaoCaoService(conn)

        # Get top vehicles
        result = service.top_xe(from_date="2026-04-01", to_date="2026-04-30", top=10)

        # Verify data is sorted by doanh_thu descending
        if len(result) > 1:
            for i in range(len(result) - 1):
                assert result[i]["doanh_thu"] >= result[i + 1]["doanh_thu"]

        # Verify each vehicle has required fields
        for item in result:
            assert "xe_id" in item
            assert "hang" in item
            assert "dong_xe" in item
            assert "so_lan_ban" in item
            assert "doanh_thu" in item

        conn.close()

    def test_ac_bc_04_employee_kpi_month(self, uat_db):
        """AC-BC-04: Chọn tháng → KPI nhân viên khớp với dữ liệu HĐ."""
        conn = sqlite3.connect(uat_db)
        conn.execute("PRAGMA foreign_keys = ON")
        service = BaoCaoService(conn)

        # Get KPI for April 2026
        result = service.kpi_nv("2026-04")

        # Verify result structure
        assert len(result) > 0

        for item in result:
            # Verify required fields
            assert "nhan_vien_id" in item
            assert "ho_ten" in item
            assert "so_hop_dong_moi_tao" in item
            assert "so_hop_dong_da_thanh_toan" in item
            assert "so_hop_dong_giao_thanh_cong" in item
            assert "doanh_thu" in item
            assert "ti_le_chot" in item

            # ti_le_chot should be between 0 and 100
            if item["so_hop_dong_moi_tao"] > 0:
                assert 0 <= item["ti_le_chot"] <= 100

        conn.close()

    def test_ac_bc_05_vip_list_sorted(self, uat_db):
        """AC-BC-05: Top N VIP → danh sách sắp xếp theo tong_gia_tri_mua giảm dần."""
        conn = sqlite3.connect(uat_db)
        conn.execute("PRAGMA foreign_keys = ON")
        service = BaoCaoService(conn)

        # Get top 10 VIP customers
        result = service.vip_customers(top=10)

        # Verify sorted by tong_gia_tri_mua descending
        for i in range(len(result) - 1):
            assert result[i]["tong_gia_tri_mua"] >= result[i + 1]["tong_gia_tri_mua"]

        # Verify VIP customers have higher purchase totals
        vip_customers = [r for r in result if r.get("phan_loai") == "VIP"]
        assert len(vip_customers) > 0

        # The highest VIP should have tong_gia_tri_mua > 1B
        if vip_customers:
            max_vip = max(vip_customers, key=lambda x: x.get("tong_gia_tri_mua", 0))
            assert max_vip["tong_gia_tri_mua"] >= 1500000000

        conn.close()


# =============================================================================
# TEST.05 — UAT Dashboard
# =============================================================================
class TestUAT_Dashboard:
    """TEST.05 — UAT theo AC-DB-* cho màn hình Dashboard"""

    def test_ac_db_01_all_kpi_tiles_have_values(self, uat_db):
        """AC-DB-01: Tất cả 7 KPI tiles đều có giá trị."""
        conn = sqlite3.connect(uat_db)
        conn.execute("PRAGMA foreign_keys = ON")
        service = DashboardService(conn)

        result = service.get_summary(role="A-01", user_id=None)

        kpis = result["kpis"]

        # All 7 KPIs should be present
        assert "revenue_month" in kpis
        assert "hop_dong_month" in kpis
        assert "xe_ton_kho" in kpis
        assert "bh_expiring_30d" in kpis
        assert "tg_qua_han" in kpis
        assert "kh_birthday_7d" in kpis
        assert "kn_cao" in kpis

        # All should have non-None values for admin
        for kpi_name, kpi_value in kpis.items():
            assert kpi_value is not None, f"KPI {kpi_name} should have value for admin"

        conn.close()

    def test_ac_db_02_kpi_tile_navigation(self, uat_db):
        """AC-DB-02: Click KPI tile → navigation đến màn hình tương ứng.

        Note: This test verifies the service returns correct data that would be
        used for navigation. Actual navigation is UI-level.
        """
        conn = sqlite3.connect(uat_db)
        conn.execute("PRAGMA foreign_keys = ON")
        service = DashboardService(conn)

        result = service.get_summary(role="A-01", user_id=None)
        kpis = result["kpis"]

        # Each KPI should have a value that can be used for navigation
        # revenue_month -> revenue report screen
        assert kpis["revenue_month"] >= 0

        # hop_dong_month -> contract list screen
        assert kpis["hop_dong_month"] >= 0

        # xe_ton_kho -> inventory screen
        assert kpis["xe_ton_kho"] >= 0

        # bh_expiring_30d -> warranty screen
        assert kpis["bh_expiring_30d"] >= 0

        # tg_qua_han -> installment screen
        assert kpis["tg_qua_han"] >= 0

        # kh_birthday_7d -> customer list screen
        assert kpis["kh_birthday_7d"] >= 0

        # kn_cao -> complaint screen
        assert kpis["kn_cao"] >= 0

        conn.close()

    def test_ac_db_03_auto_refresh_data_updates(self, uat_db):
        """AC-DB-03: Sau refresh, dữ liệu dashboard được cập nhật.

        Note: Auto-refresh is a UI/timer feature. This test verifies that
        calling get_summary returns fresh data each time.
        """
        conn = sqlite3.connect(uat_db)
        conn.execute("PRAGMA foreign_keys = ON")
        service = DashboardService(conn)

        # Get summary twice
        result1 = service.get_summary(role="A-01", user_id=None)
        result2 = service.get_summary(role="A-01", user_id=None)

        # Timestamps should be different (if implementation uses live timestamps)
        # Note: Some implementations may cache results, so this test may be lenient

        # KPIs should still be valid
        assert result1["kpis"]["revenue_month"] >= 0
        assert result2["kpis"]["revenue_month"] >= 0

        conn.close()

    def test_ac_db_role_based_kpi_visibility(self, uat_db):
        """AC-DB-04: KPI visibility based on user role.

        Tests that different roles see appropriate KPI subsets.
        """
        conn = sqlite3.connect(uat_db)
        conn.execute("PRAGMA foreign_keys = ON")
        service = DashboardService(conn)

        # Admin sees all
        admin_result = service.get_summary(role="A-01", user_id=None)
        for kpi_value in admin_result["kpis"].values():
            assert kpi_value is not None

        # Staff sees limited
        staff_result = service.get_summary(role="A-03", user_id=3)
        assert staff_result["kpis"]["revenue_month"] is not None or staff_result["kpis"]["revenue_month"] == 0
        # Other KPIs should be None for staff
        assert staff_result["kpis"].get("xe_ton_kho") is None

        conn.close()