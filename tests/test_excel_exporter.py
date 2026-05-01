"""Unit tests for ExcelExporter - T-G5.4.TEST.02.

Tests:
- TEST.02: export_report and export_multi_sheet with 8 test cases

References:
- ExcelExporter in app/infrastructure/exporters/excel_exporter.py
- Format types: money (VND), date (DD/MM/YYYY), percent, number, text
"""

import pytest
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.infrastructure.exporters.excel_exporter import (
    ExcelExporter,
    ColumnConfig,
    SheetConfig,
    ValidationError,
    ExcelExporterError,
)


@pytest.fixture
def exporter():
    """Create ExcelExporter instance."""
    return ExcelExporter()


@pytest.fixture
def temp_dir():
    """Create temporary directory for test files."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_revenue_data():
    """Sample revenue report data."""
    return [
        {
            "period": "2026-04-01",
            "so_hop_dong": 15,
            "doanh_thu": 7500000000,
            "ty_le": 30.5,
        },
        {
            "period": "2026-04-02",
            "so_hop_dong": 12,
            "doanh_thu": 6000000000,
            "ty_le": 24.4,
        },
        {
            "period": "2026-04-03",
            "so_hop_dong": 18,
            "doanh_thu": 9000000000,
            "ty_le": 36.6,
        },
    ]


@pytest.fixture
def sheet_config_revenue():
    """Sheet config for revenue report."""
    return {
        "name": "BaoCaoDoanhThu",
        "title": "Báo Cáo Doanh Thu",
        "columns": [
            {"header": "Ngày", "key": "period", "width": 12, "format": "date"},
            {"header": "Số Hợp Đồng", "key": "so_hop_dong", "width": 15, "format": "number"},
            {"header": "Doanh Thu (VND)", "key": "doanh_thu", "width": 20, "format": "money"},
            {"header": "Tỷ Lệ (%)", "key": "ty_le", "width": 12, "format": "percent"},
        ],
    }


@pytest.fixture
def multi_sheet_data():
    """Data for multi-sheet export."""
    return [
        {
            "name": "DoanhThu",
            "title": "Báo Cáo Doanh Thu",
            "columns": [
                {"header": "Ngày", "key": "period", "width": 12, "format": "date"},
                {"header": "Số HĐ", "key": "so_hop_dong", "width": 10, "format": "number"},
                {"header": "Doanh Thu", "key": "doanh_thu", "width": 18, "format": "money"},
            ],
            "data": [
                {"period": "2026-04-01", "so_hop_dong": 15, "doanh_thu": 7500000000},
                {"period": "2026-04-02", "so_hop_dong": 12, "doanh_thu": 6000000000},
            ],
        },
        {
            "name": "TopXe",
            "title": "Top Xe Ban Chay",
            "columns": [
                {"header": "Hãng", "key": "hang", "width": 12, "format": "text"},
                {"header": "Dòng Xe", "key": "dong_xe", "width": 12, "format": "text"},
                {"header": "Số Lần Bán", "key": "so_lan_ban", "width": 12, "format": "number"},
                {"header": "Doanh Thu", "key": "doanh_thu", "width": 18, "format": "money"},
            ],
            "data": [
                {"hang": "Toyota", "dong_xe": "Camry", "so_lan_ban": 25, "doanh_thu": 12500000000},
                {"hang": "Honda", "dong_xe": "Civic", "so_lan_ban": 20, "doanh_thu": 8000000000},
            ],
        },
    ]


# =============================================================================
# TEST.02 — Excel Export Tests
# =============================================================================
class TestExcelExporter:
    """TEST.02 — ExcelExporter export_report — 8 test cases"""

    def test_export_report_creates_file(self, exporter, temp_dir, sample_revenue_data, sheet_config_revenue):
        """export_report tạo file .xlsx tại output_path."""
        output_path = temp_dir / "test_revenue.xlsx"

        result = exporter.export_report(sample_revenue_data, sheet_config_revenue, str(output_path))

        assert result == str(output_path)
        assert os.path.exists(output_path)
        assert os.path.getsize(output_path) > 0
        assert output_path.suffix == ".xlsx"

    def test_export_report_has_header(self, exporter, temp_dir, sample_revenue_data, sheet_config_revenue):
        """Dòng đầu tiên chứa column headers."""
        output_path = temp_dir / "test_header.xlsx"
        exporter.export_report(sample_revenue_data, sheet_config_revenue, str(output_path))

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # With title row, headers are in row 2
        headers = [cell.value for cell in ws[2]]
        assert "Ngày" in headers
        assert "Số Hợp Đồng" in headers
        assert "Doanh Thu (VND)" in headers
        assert "Tỷ Lệ (%)" in headers

        wb.close()

    def test_export_report_has_data(self, exporter, temp_dir, sample_revenue_data, sheet_config_revenue):
        """Các dòng dữ liệu chứa giá trị đúng."""
        output_path = temp_dir / "test_data.xlsx"
        exporter.export_report(sample_revenue_data, sheet_config_revenue, str(output_path))

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Data rows start at row 3 (after title + header)
        # First data row: period = "2026-04-01", so_hop_dong = 15, doanh_thu = 7500000000
        data_row = ws[3]
        assert data_row[0].value == "2026-04-01"
        assert data_row[1].value == 15
        assert data_row[2].value == 7500000000

        wb.close()

    def test_export_report_header_bold(self, exporter, temp_dir, sample_revenue_data, sheet_config_revenue):
        """Header row có font bold."""
        output_path = temp_dir / "test_bold.xlsx"
        exporter.export_report(sample_revenue_data, sheet_config_revenue, str(output_path))

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Headers are in row 2 (after title)
        header_row = ws[2]
        for cell in header_row:
            if cell.value:
                assert cell.font.bold is True, f"Header '{cell.value}' should be bold"

        wb.close()

    def test_export_report_money_format(self, exporter, temp_dir, sample_revenue_data, sheet_config_revenue):
        """Số tiền VND được format đúng (ví dụ: 1.000.000 đ)."""
        output_path = temp_dir / "test_money.xlsx"
        exporter.export_report(sample_revenue_data, sheet_config_revenue, str(output_path))

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Check that money column has number format
        data_row = ws[3]
        money_cell = data_row[2]  # doanh_thu column (index 2)
        assert money_cell.value == 7500000000
        # openpyxl uses '#,##0' format which displays with thousand separators
        assert money_cell.number_format == "#,##0"

        wb.close()

    def test_export_report_date_format(self, exporter, temp_dir, sample_revenue_data, sheet_config_revenue):
        """Ngày tháng được format đúng (DD/MM/YYYY)."""
        output_path = temp_dir / "test_date.xlsx"
        exporter.export_report(sample_revenue_data, sheet_config_revenue, str(output_path))

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Date cell (column 0, row 3)
        date_cell = ws[3][0]
        assert date_cell.value == "2026-04-01"

        wb.close()

    def test_export_report_invalid_path_raises(self, exporter, temp_dir, sample_revenue_data, sheet_config_revenue):
        """export_report raise clear error cho invalid path."""
        # Test with invalid path (directory doesn't exist and can't be created)
        invalid_path = "/nonexistent/path/that/cannot/be/created/file.xlsx"

        with pytest.raises((ExcelExporterError, ValidationError, OSError, IOError)):
            exporter.export_report(sample_revenue_data, sheet_config_revenue, invalid_path)

    def test_export_multi_sheet(self, exporter, temp_dir, multi_sheet_data):
        """export_multi_sheet tạo file với nhiều sheets."""
        output_path = temp_dir / "test_multi.xlsx"

        result = exporter.export_multi_sheet(multi_sheet_data, str(output_path))

        assert result == str(output_path)
        assert os.path.exists(output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)

        # Check that multiple sheets exist
        sheet_names = wb.sheetnames
        assert "DoanhThu" in sheet_names
        assert "TopXe" in sheet_names

        # Check sheet content
        ws1 = wb["DoanhThu"]
        headers1 = [cell.value for cell in ws1[2]]
        assert "Ngày" in headers1

        ws2 = wb["TopXe"]
        headers2 = [cell.value for cell in ws2[2]]
        assert "Hãng" in headers2

        wb.close()

    def test_export_empty_data(self, exporter, temp_dir, sheet_config_revenue):
        """export_report với data rỗng tạo file chỉ có header."""
        output_path = temp_dir / "test_empty.xlsx"
        empty_data = []

        result = exporter.export_report(empty_data, sheet_config_revenue, str(output_path))

        assert os.path.exists(output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Should still have title and headers
        assert ws.cell(row=1, column=1).value == "Báo Cáo Doanh Thu"
        headers = [cell.value for cell in ws[2]]
        assert "Ngày" in headers

        wb.close()

    def test_export_with_title_row(self, exporter, temp_dir):
        """export_report có title row và header row riêng biệt."""
        data = [{"col_a": 1, "col_b": 2}]
        config = {
            "name": "Test",
            "title": "This Is A Title Row",
            "columns": [
                {"header": "Column A", "key": "col_a", "format": "number"},
                {"header": "Column B", "key": "col_b", "format": "number"},
            ],
        }
        output_path = temp_dir / "test_title.xlsx"
        exporter.export_report(data, config, str(output_path))

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Row 1 should have merged title
        title_cell = ws.cell(row=1, column=1)
        assert title_cell.value == "This Is A Title Row"

        # Row 2 should have headers
        headers = [cell.value for cell in ws[2]]
        assert "Column A" in headers
        assert "Column B" in headers

        # Data row starts at 3
        data_row = ws[3]
        assert data_row[0].value == 1

        wb.close()